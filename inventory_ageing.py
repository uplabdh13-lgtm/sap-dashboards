"""
=============================================================
  SAP Inventory Ageing Dashboard
  Author: Generated for SAP MM Consultant
  Sources: MB52 (Stock on Hand) + MB51 (Material Movements)
  Output:  1) inventory_ageing_report.xlsx
           2) inventory_ageing_dashboard.html
=============================================================

HOW TO USE:
-----------
1. In SAP, run MB52 → Export to Excel → Save as 'mb52_export.xlsx'
2. In SAP, run MB51 → Export to Excel → Save as 'mb51_export.xlsx'
3. Place both files in the same folder as this script
4. Run: python inventory_ageing.py
5. Open inventory_ageing_dashboard.html in your browser

EXPECTED COLUMN NAMES (adjust MB52_COLS and MB51_COLS below if different):
---------------------------------------------------------------------------
MB52: Material, Material Description, Plant, Storage Location,
      Batch, Unrestricted, Total Stock Value
MB51: Material, Plant, Storage Location, Batch,
      Posting Date, Movement Type
=============================================================
"""

import pandas as pd
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime, date
import json, os, sys

# ─────────────────────────────────────────────
#  CONFIGURATION  ← Edit here if needed
# ─────────────────────────────────────────────
# Auto-detect file location (local or GitHub Actions)
import os as _os
MB52_FILE = "data/mb52_export.xlsx" if _os.path.exists("data/mb52_export.xlsx") else "mb52_export.xlsx"
MB51_FILE = "data/mb51_export.xlsx" if _os.path.exists("data/mb51_export.xlsx") else "mb51_export.xlsx"

# Ageing buckets (days)
BUCKETS = [
    (0,   30,  "0-30 Days",   "#22c55e"),
    (31,  90,  "31-90 Days",  "#84cc16"),
    (91,  180, "91-180 Days", "#eab308"),
    (181, 360, "181-360 Days","#f97316"),
    (361, 9999,"360+ Days",   "#ef4444"),
]

# Column name mappings — matched to your actual SAP export headers
MB52_COLS = {
    "material":     "Material",
    "description":  "Material",           # MB52 has no description col; use Material as fallback
    "plant":        "Plant",
    "sloc":         "Storage Location",
    "batch":        "",                    # MB52 has no Batch column
    "qty":          "Unrestricted",
    "value":        "Value Unrestricted",  # actual column name in your MB52
}

MB51_COLS = {
    "material":     "Material",
    "plant":        "Plant",
    "sloc":         "Storage Location",
    "batch":        "Batch",
    "post_date":    "Posting Date",
    "mvt_type":     "Movement Type",
}

OUTPUT_EXCEL = "inventory_ageing_report.xlsx"
OUTPUT_HTML  = "inventory_ageing_dashboard.html"
TODAY        = date.today()

# ─────────────────────────────────────────────
#  HELPER: Assign ageing bucket
# ─────────────────────────────────────────────
def get_bucket(days):
    for lo, hi, label, color in BUCKETS:
        if lo <= days <= hi:
            return label, color
    return "360+ Days", "#ef4444"

# ─────────────────────────────────────────────
#  STEP 1: Load Data
# ─────────────────────────────────────────────
def load_data():
    print("📂 Loading MB52 and MB51 data...")

    # ── Demo / Sample data when real files are missing ──────────────
    # Also copy uploaded files if available
    import shutil
    for src, dst in [
        ("/mnt/user-data/uploads/mb52_export.xlsx", MB52_FILE),
        ("/mnt/user-data/uploads/mb51_export.xlsx", MB51_FILE),
    ]:
        if os.path.exists(src) and not os.path.exists(dst):
            shutil.copy(src, dst)
            print(f"   Copied {src} → {dst}")

    if not os.path.exists(MB52_FILE) or not os.path.exists(MB51_FILE):
        print("⚠️  Export files not found — generating SAMPLE data for demo.")
        import random
        random.seed(42)
        materials = [
            ("MAT-001","Hydraulic Oil 10L","1000","SL01","BATCH-A"),
            ("MAT-002","Packing Carton","1000","SL01","BATCH-B"),
            ("MAT-003","Spare Motor Part","2000","SL02","BATCH-C"),
            ("MAT-004","Lubricant Grease","1000","SL03","BATCH-D"),
            ("MAT-005","Steel Rod 6mm","2000","SL02","BATCH-E"),
            ("MAT-006","Filter Element","1000","SL01","BATCH-F"),
            ("MAT-007","Gasket Set","3000","SL04","BATCH-G"),
            ("MAT-008","Pump Seal","2000","SL02","BATCH-H"),
            ("MAT-009","Electrical Cable","1000","SL03","BATCH-I"),
            ("MAT-010","Bearing 6205","3000","SL04","BATCH-J"),
            ("MAT-011","Bolt M12","1000","SL01","BATCH-K"),
            ("MAT-012","Paint - White 20L","2000","SL02","BATCH-L"),
            ("MAT-013","Safety Gloves","3000","SL04","BATCH-M"),
            ("MAT-014","Hydraulic Hose","1000","SL01","BATCH-N"),
            ("MAT-015","Valve 2 inch","2000","SL02","BATCH-O"),
        ]
        mb52_rows = []
        mb51_rows = []
        for mat, desc, plant, sloc, batch in materials:
            qty   = random.randint(10, 500)
            value = round(qty * random.uniform(20, 500), 2)
            mb52_rows.append({
                MB52_COLS["material"]:    mat,
                MB52_COLS["description"]: desc,
                MB52_COLS["plant"]:       plant,
                MB52_COLS["sloc"]:        sloc,
                MB52_COLS["batch"]:       batch,
                MB52_COLS["qty"]:         qty,
                MB52_COLS["value"]:       value,
            })
            # Last movement between 5 and 450 days ago
            days_ago = random.randint(5, 450)
            post_date = pd.Timestamp(TODAY) - pd.Timedelta(days=days_ago)
            mb51_rows.append({
                MB51_COLS["material"]:   mat,
                MB51_COLS["plant"]:      plant,
                MB51_COLS["sloc"]:       sloc,
                MB51_COLS["batch"]:      batch,
                MB51_COLS["post_date"]:  post_date,
                MB51_COLS["mvt_type"]:   random.choice(["101","201","261","301"]),
            })

        mb52 = pd.DataFrame(mb52_rows)
        mb51 = pd.DataFrame(mb51_rows)
    else:
        mb52 = pd.read_excel(MB52_FILE)
        mb51 = pd.read_excel(MB51_FILE)

    return mb52, mb51

# ─────────────────────────────────────────────
#  STEP 2: Calculate Ageing
# ─────────────────────────────────────────────
def calculate_ageing(mb52, mb51):
    print("⚙️  Calculating inventory ageing...")

    c  = MB51_COLS
    m52 = MB52_COLS

    mb51[c["post_date"]] = pd.to_datetime(mb51[c["post_date"]], errors="coerce")

    # Fill NaN in key columns so groupby/merge don't drop rows
    for col in [c["material"], c["plant"], c["sloc"], c["batch"]]:
        if col and col in mb51.columns:
            mb51[col] = mb51[col].fillna("").astype(str).str.strip()
    for col in [m52["material"], m52["plant"], m52["sloc"]]:
        if col and col in mb52.columns:
            mb52[col] = mb52[col].fillna("").astype(str).str.strip()

    # Add Description column if missing in MB52
    if "Material Description" not in mb52.columns:
        mb52["Material Description"] = mb52.get(m52["material"], "")

    # Always merge on Material + Plant only (most reliable across MB52/MB51)
    group_keys_mb51 = [c["material"], c["plant"]]
    group_keys_mb52 = [m52["material"], m52["plant"]]

    # Include Sloc only if BOTH files have it with real data
    sloc_in_mb51 = c["sloc"] in mb51.columns and mb51[c["sloc"]].replace("",pd.NA).dropna().shape[0] > 0
    sloc_in_mb52 = m52["sloc"] in mb52.columns and mb52[m52["sloc"]].replace("",pd.NA).dropna().shape[0] > 0
    if sloc_in_mb51 and sloc_in_mb52:
        group_keys_mb51.append(c["sloc"])
        group_keys_mb52.append(m52["sloc"])

    # Skip Batch merge — MB52 has no Batch column

    print(f"   Merging on keys: {group_keys_mb51}")

    last_mvt = (
        mb51.groupby(group_keys_mb51)[c["post_date"]]
        .max()
        .reset_index()
        .rename(columns={c["post_date"]: "Last_Movement_Date"})
    )

    # Merge with MB52
    df = mb52.merge(
        last_mvt,
        left_on  =group_keys_mb52,
        right_on =group_keys_mb51,
        how="left",
        suffixes=("","_mb51")
    )

    # Drop duplicate key columns from MB51 side
    for col in group_keys_mb51:
        if col+"_mb51" in df.columns:
            df.drop(columns=[col+"_mb51"], inplace=True)

    # Days since last movement
    df["Last_Movement_Date"] = pd.to_datetime(df["Last_Movement_Date"])
    df["Days_Since_Movement"] = (
        pd.Timestamp(TODAY) - df["Last_Movement_Date"]
    ).dt.days.fillna(9999).astype(int)

    # Assign bucket
    df["Ageing_Bucket"] = df["Days_Since_Movement"].apply(lambda d: get_bucket(d)[0])
    df["Bucket_Color"]  = df["Days_Since_Movement"].apply(lambda d: get_bucket(d)[1])

    # Clean up
    df["Last_Movement_Date"] = df["Last_Movement_Date"].dt.strftime("%Y-%m-%d").fillna("No Movement")

    return df

# ─────────────────────────────────────────────
#  STEP 3: Export to Excel
# ─────────────────────────────────────────────
def export_excel(df):
    print("📊 Generating Excel report...")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Detail ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Inventory Ageing Detail"

    COLUMNS = [
        (MB52_COLS["material"],    "Material",          18),
        ("Material Description",   "Description",       30),
        (MB52_COLS["plant"],       "Plant",             10),
        (MB52_COLS["sloc"],        "Storage Loc",       14),
        (MB51_COLS["batch"],       "Batch",             14),
        (MB52_COLS["qty"],         "Qty (Unrestricted)",18),
        (MB52_COLS["value"],       "Stock Value (AED)", 18),
        ("Last_Movement_Date",     "Last Movement",     16),
        ("Days_Since_Movement",    "Days",              10),
        ("Ageing_Bucket",          "Ageing Bucket",     16),
    ]

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1e3a5f")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 28
    for col_idx, (_, header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Bucket fill map
    bucket_fills = {label: PatternFill("solid", fgColor=color.lstrip("#"))
                    for _, _, label, color in BUCKETS}

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        bucket = row.get("Ageing_Bucket", "0-30 Days")
        for col_idx, (col_key, _, _) in enumerate(COLUMNS, 1):
            val  = row.get(col_key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(size=10)
            # Colour the Ageing Bucket column
            if col_key == "Ageing_Bucket":
                cell.fill = bucket_fills.get(bucket, PatternFill())
                cell.font = Font(bold=True, size=10)

        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(COLUMNS)):
                c2 = ws.cell(row=row_idx, column=col_idx)
                if not c2.fill or c2.fill.fgColor.rgb in ("00000000","FFFFFFFF"):
                    c2.fill = PatternFill("solid", fgColor="F0F4F8")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Ageing Summary")
    ws2.sheet_view.showGridLines = False

    # Title
    ws2.merge_cells("A1:E1")
    title_cell = ws2["A1"]
    title_cell.value     = f"Inventory Ageing Summary  |  As of {TODAY.strftime('%d %B %Y')}"
    title_cell.font      = Font(bold=True, size=14, color="1e3a5f")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill      = PatternFill("solid", fgColor="E8F0FE")
    ws2.row_dimensions[1].height = 36

    headers2 = ["Ageing Bucket","No. of Materials","Total Qty","Total Value (AED)","% of Total Value"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = 22

    summary = (
        df.groupby("Ageing_Bucket")
        .agg(
            Materials =(MB52_COLS["material"], "count"),
            Total_Qty =(MB52_COLS["qty"],      "sum"),
            Total_Value=(MB52_COLS["value"],    "sum"),
        )
        .reset_index()
    )
    total_val = summary["Total_Value"].sum()
    bucket_order = [b[2] for b in BUCKETS]
    summary["sort_key"] = summary["Ageing_Bucket"].apply(
        lambda x: bucket_order.index(x) if x in bucket_order else 99)
    summary = summary.sort_values("sort_key")

    for row_idx, srow in enumerate(summary.itertuples(), 4):
        pct = round((srow.Total_Value / total_val * 100), 1) if total_val else 0
        vals = [srow.Ageing_Bucket, srow.Materials,
                round(srow.Total_Qty, 2),
                round(srow.Total_Value, 2), f"{pct}%"]
        fill_color = next((c.lstrip("#") for _,_,l,c in BUCKETS
                           if l == srow.Ageing_Bucket), "FFFFFF")
        for col_idx, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=v)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(size=11)
            if col_idx == 1:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=True, size=11)

    # Totals row
    tot_row = 4 + len(summary)
    tot_vals = ["TOTAL", summary["Materials"].sum(),
                round(summary["Total_Qty"].sum(), 2),
                round(total_val, 2), "100%"]
    for col_idx, v in enumerate(tot_vals, 1):
        cell = ws2.cell(row=tot_row, column=col_idx, value=v)
        cell.fill      = PatternFill("solid", fgColor="1e3a5f")
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    # Bar chart
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Stock Value by Ageing Bucket"
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Bucket"
    chart.width   = 20
    chart.height  = 12
    data_ref  = Reference(ws2, min_col=4, min_row=3,
                          max_row=3+len(summary))
    cats_ref  = Reference(ws2, min_col=1, min_row=4,
                          max_row=3+len(summary))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, "G3")

    wb.save(OUTPUT_EXCEL)
    print(f"   ✅ Excel saved → {OUTPUT_EXCEL}")
    return summary, total_val

# ─────────────────────────────────────────────
#  STEP 4: Export to HTML Dashboard
# ─────────────────────────────────────────────
def export_html(df, summary, total_val):
    print("🌐 Generating HTML dashboard...")

    bucket_order = [b[2] for b in BUCKETS]
    summary["sort_key"] = summary["Ageing_Bucket"].apply(
        lambda x: bucket_order.index(x) if x in bucket_order else 99)
    summary = summary.sort_values("sort_key")

    bucket_labels  = summary["Ageing_Bucket"].tolist()
    bucket_values  = summary["Total_Value"].round(2).tolist()
    bucket_counts  = summary["Materials"].tolist()
    bucket_colors  = [next((c for _,_,l,c in BUCKETS if l==b), "#999")
                      for b in bucket_labels]

    total_materials = int(df[MB52_COLS["material"]].count())
    total_qty       = round(float(df[MB52_COLS["qty"]].sum()), 2)
    slow_moving_val = round(float(
        df[df["Days_Since_Movement"] > 180][MB52_COLS["value"]].sum()), 2)
    slow_pct        = round(slow_moving_val / total_val * 100, 1) if total_val else 0

    # Table rows
    table_rows_html = ""
    for _, row in df.iterrows():
        bucket = row.get("Ageing_Bucket","")
        color  = row.get("Bucket_Color","#999")
        table_rows_html += f"""
        <tr>
          <td>{row.get(MB52_COLS['material'],'')}</td>
          <td>{row.get(MB52_COLS['description'],'')}</td>
          <td>{row.get(MB52_COLS['plant'],'')}</td>
          <td>{row.get(MB52_COLS['sloc'],'')}</td>
          <td>{row.get(MB52_COLS['batch'],'')}</td>
          <td>{row.get(MB52_COLS['qty'],'')}</td>
          <td>{row.get(MB52_COLS['value'],'')}</td>
          <td>{row.get('Last_Movement_Date','')}</td>
          <td>{row.get('Days_Since_Movement','')}</td>
          <td><span class="badge" style="background:{color}">{bucket}</span></td>
        </tr>"""

    summary_rows_html = ""
    for _, srow in summary.iterrows():
        pct   = round(srow["Total_Value"] / total_val * 100, 1) if total_val else 0
        color = next((c for _,_,l,c in BUCKETS if l==srow["Ageing_Bucket"]), "#999")
        summary_rows_html += f"""
        <tr>
          <td><span class="badge" style="background:{color}">{srow['Ageing_Bucket']}</span></td>
          <td>{int(srow['Materials'])}</td>
          <td>{round(srow['Total_Qty'],2):,}</td>
          <td>{round(srow['Total_Value'],2):,.2f}</td>
          <td>
            <div class="progress-bar">
              <div class="progress-fill" style="width:{pct}%;background:{color}"></div>
            </div>
            <span style="font-size:11px;color:#64748b">{pct}%</span>
          </td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Inventory Ageing Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  :root {{
    --navy:   #0f2344;
    --blue:   #1e3a5f;
    --accent: #3b82f6;
    --bg:     #f1f5f9;
    --card:   #ffffff;
    --text:   #1e293b;
    --muted:  #64748b;
  }}
  body {{ font-family:'DM Sans',sans-serif; background:var(--bg); color:var(--text); }}

  /* ── Header ── */
  .header {{
    background: linear-gradient(135deg, var(--navy) 0%, #1a4080 100%);
    padding: 28px 40px; display:flex; align-items:center;
    justify-content:space-between; box-shadow:0 4px 24px rgba(0,0,0,.25);
  }}
  .header-left h1 {{ font-size:22px; font-weight:700; color:#fff; letter-spacing:-.3px; }}
  .header-left p  {{ font-size:13px; color:#93c5fd; margin-top:3px; font-family:'DM Mono',monospace; }}
  .header-badge {{
    background:rgba(255,255,255,.1); border:1px solid rgba(255,255,255,.2);
    color:#e0f2fe; font-size:12px; padding:6px 14px; border-radius:20px;
    font-family:'DM Mono',monospace;
  }}

  /* ── Layout ── */
  .container {{ max-width:1400px; margin:0 auto; padding:32px 24px; }}

  /* ── KPI Cards ── */
  .kpi-grid {{ display:grid; grid-template-columns:repeat(4,1fr); gap:18px; margin-bottom:28px; }}
  .kpi-card {{
    background:var(--card); border-radius:14px; padding:22px 24px;
    box-shadow:0 1px 4px rgba(0,0,0,.07); border-left:4px solid var(--accent);
    transition:transform .2s;
  }}
  .kpi-card:hover {{ transform:translateY(-3px); box-shadow:0 6px 20px rgba(0,0,0,.1); }}
  .kpi-card .label {{ font-size:12px; font-weight:600; color:var(--muted); text-transform:uppercase; letter-spacing:.6px; }}
  .kpi-card .value {{ font-size:30px; font-weight:700; color:var(--navy); margin:6px 0 2px; line-height:1; }}
  .kpi-card .sub   {{ font-size:12px; color:var(--muted); }}
  .kpi-card.warn   {{ border-left-color:#ef4444; }}
  .kpi-card.warn .value {{ color:#dc2626; }}

  /* ── Charts ── */
  .charts-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:18px; margin-bottom:28px; }}
  .chart-card {{
    background:var(--card); border-radius:14px; padding:24px;
    box-shadow:0 1px 4px rgba(0,0,0,.07);
  }}
  .chart-card h3 {{ font-size:14px; font-weight:600; color:var(--navy); margin-bottom:18px; }}
  .chart-wrap {{ position:relative; height:260px; }}

  /* ── Summary Table ── */
  .section-title {{
    font-size:15px; font-weight:700; color:var(--navy);
    margin-bottom:14px; display:flex; align-items:center; gap:8px;
  }}
  .section-title::before {{
    content:''; display:block; width:4px; height:18px;
    background:var(--accent); border-radius:2px;
  }}
  .table-card {{
    background:var(--card); border-radius:14px; padding:24px;
    box-shadow:0 1px 4px rgba(0,0,0,.07); margin-bottom:28px; overflow-x:auto;
  }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  th {{
    background:var(--navy); color:#fff; font-weight:600; font-size:12px;
    padding:11px 14px; text-align:left; white-space:nowrap;
  }}
  th:first-child {{ border-radius:6px 0 0 6px; }}
  th:last-child  {{ border-radius:0 6px 6px 0; }}
  td {{ padding:10px 14px; border-bottom:1px solid #e2e8f0; vertical-align:middle; }}
  tr:hover td {{ background:#f8faff; }}
  .badge {{
    display:inline-block; padding:4px 11px; border-radius:20px;
    font-size:11px; font-weight:600; color:#fff; white-space:nowrap;
  }}
  .progress-bar {{ background:#e2e8f0; border-radius:6px; height:8px; width:120px; display:inline-block; vertical-align:middle; margin-right:6px; }}
  .progress-fill {{ height:100%; border-radius:6px; transition:width .5s; }}

  /* ── Filter bar ── */
  .filter-bar {{ display:flex; gap:12px; flex-wrap:wrap; margin-bottom:16px; align-items:center; }}
  .filter-bar input, .filter-bar select {{
    padding:8px 14px; border:1px solid #cbd5e1; border-radius:8px;
    font-family:'DM Sans',sans-serif; font-size:13px; color:var(--text);
    background:var(--card); outline:none;
  }}
  .filter-bar input:focus, .filter-bar select:focus {{ border-color:var(--accent); }}
  .filter-bar label {{ font-size:12px; font-weight:600; color:var(--muted); text-transform:uppercase; letter-spacing:.5px; }}

  /* ── Footer ── */
  .footer {{ text-align:center; padding:20px; font-size:12px; color:var(--muted); }}
</style>
</head>
<body>

<div class="header">
  <div class="header-left">
    <h1>📦 Inventory Ageing Dashboard</h1>
    <p>MB52 Stock on Hand  ·  MB51 Last Movement  ·  Generated {TODAY.strftime('%d %B %Y')}</p>
  </div>
  <div class="header-badge">SAP MM · Ageing Analysis</div>
</div>

<div class="container">

  <!-- KPI Cards -->
  <div class="kpi-grid">
    <div class="kpi-card">
      <div class="label">Total Materials</div>
      <div class="value">{total_materials:,}</div>
      <div class="sub">Unique material records</div>
    </div>
    <div class="kpi-card">
      <div class="label">Total Stock Qty</div>
      <div class="value">{total_qty:,.0f}</div>
      <div class="sub">Unrestricted stock</div>
    </div>
    <div class="kpi-card">
      <div class="label">Total Stock Value</div>
      <div class="value">{total_val:,.0f}</div>
      <div class="sub">All buckets combined</div>
    </div>
    <div class="kpi-card warn">
      <div class="label">Slow Moving (180+ days)</div>
      <div class="value">{slow_moving_val:,.0f}</div>
      <div class="sub">{slow_pct}% of total value</div>
    </div>
  </div>

  <!-- Charts -->
  <div class="charts-grid">
    <div class="chart-card">
      <h3>📊 Stock Value by Ageing Bucket</h3>
      <div class="chart-wrap"><canvas id="barChart"></canvas></div>
    </div>
    <div class="chart-card">
      <h3>🥧 Materials Distribution</h3>
      <div class="chart-wrap"><canvas id="pieChart"></canvas></div>
    </div>
  </div>

  <!-- Summary Table -->
  <div class="table-card">
    <div class="section-title">Ageing Summary by Bucket</div>
    <table>
      <thead>
        <tr>
          <th>Ageing Bucket</th><th>No. of Materials</th>
          <th>Total Qty</th><th>Total Value</th><th>% of Total Value</th>
        </tr>
      </thead>
      <tbody>{summary_rows_html}</tbody>
    </table>
  </div>

  <!-- Detail Table -->
  <div class="table-card">
    <div class="section-title">Inventory Detail</div>
    <div class="filter-bar">
      <label>Search:</label>
      <input type="text" id="searchInput" placeholder="Material, Description, Plant..." oninput="filterTable()">
      <label>Bucket:</label>
      <select id="bucketFilter" onchange="filterTable()">
        <option value="">All Buckets</option>
        {''.join(f'<option value="{b[2]}">{b[2]}</option>' for b in BUCKETS)}
      </select>
    </div>
    <div style="overflow-x:auto">
    <table id="detailTable">
      <thead>
        <tr>
          <th>Material</th><th>Description</th><th>Plant</th>
          <th>Stor. Loc</th><th>Batch</th><th>Qty</th>
          <th>Value</th><th>Last Movement</th><th>Days</th><th>Ageing Bucket</th>
        </tr>
      </thead>
      <tbody id="detailBody">{table_rows_html}</tbody>
    </table>
    </div>
  </div>

</div>

<div class="footer">Generated by SAP Inventory Ageing Python Script · {TODAY.strftime('%d %B %Y')}</div>

<script>
// ── Charts ────────────────────────────────────────────────
const labels  = {json.dumps(bucket_labels)};
const values  = {json.dumps(bucket_values)};
const counts  = {json.dumps(bucket_counts)};
const colors  = {json.dumps(bucket_colors)};

new Chart(document.getElementById('barChart'), {{
  type: 'bar',
  data: {{
    labels,
    datasets: [{{ label:'Stock Value', data: values,
      backgroundColor: colors, borderRadius: 6, borderSkipped: false }}]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      y: {{ grid: {{ color:'#e2e8f0' }}, ticks: {{ font: {{ size:11 }} }} }},
      x: {{ grid: {{ display:false }}, ticks: {{ font: {{ size:11 }} }} }}
    }}
  }}
}});

new Chart(document.getElementById('pieChart'), {{
  type: 'doughnut',
  data: {{
    labels,
    datasets: [{{ data: counts, backgroundColor: colors,
      borderWidth: 2, borderColor:'#fff' }}]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{
      legend: {{ position:'right', labels: {{ font: {{ size:11 }} }} }}
    }}
  }}
}});

// ── Filter ────────────────────────────────────────────────
function filterTable() {{
  const search = document.getElementById('searchInput').value.toLowerCase();
  const bucket = document.getElementById('bucketFilter').value.toLowerCase();
  const rows   = document.querySelectorAll('#detailBody tr');
  rows.forEach(row => {{
    const text   = row.textContent.toLowerCase();
    const matchS = !search || text.includes(search);
    const matchB = !bucket || text.includes(bucket);
    row.style.display = (matchS && matchB) ? '' : 'none';
  }});
}}
</script>
</body>
</html>"""

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"   ✅ HTML saved → {OUTPUT_HTML}")

# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  SAP Inventory Ageing Dashboard Generator")
    print(f"  Run Date: {TODAY.strftime('%d %B %Y')}")
    print("=" * 55)

    mb52, mb51   = load_data()
    df           = calculate_ageing(mb52, mb51)
    summary, tv  = export_excel(df)
    export_html(df, summary, tv)

    print()
    print("✅ DONE!")
    print(f"   📊 Excel  → {OUTPUT_EXCEL}")
    print(f"   🌐 HTML   → {OUTPUT_HTML}")
    print()
    print("   Open the HTML file in your browser for the interactive dashboard.")
    print("=" * 55)
